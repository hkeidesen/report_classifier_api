from pdfminer3.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer3.converter import TextConverter
from pdfminer3.layout import LAParams
from pdfminer3.pdfpage import PDFPage

from io import StringIO
from io import BytesIO

import re
import requests
import json 
import openpyxl
from urllib.request import urlopen, Request
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from time import sleep

from Installation_Type_Dictionary import Installations

class Report:
    
    def __init__(self, year, date, activity_number, title, taskleader,
                 participants_in_revision, count_participants, url, installation_name, installation_type):
        #self.authority = "PTIL"'
        self.year = year
        self.date = date
        self.activity_number = activity_number
        self.title = title
        self.taskleader = taskleader
        self.participants_in_revision = participants_in_revision
        self.count_participants = count_participants
        self.url = url
        self.installation_name = installation_name
        self.installation_type = installation_type
        
        self.deviation_list = []
        self.improvement_list = []

class Deviation:

    def __init__(self, title, description, regulations):
        self.finding_type = "Avvik"
        self.title = title
        self.description = description
        self.regulations = regulations
        #self.cathegory = cathegory

class Improvementpoint:
    
    def __init__(self, title, description, regulations):
        self.finding_type = "Forbedringspunkt"
        self.title = title
        self.description = description
        self.regulations = regulations
        #self.cathegory = cathegory
    
## Make soup of webpage
def make_soup(url):
    
    res = requests.get(url)
    soup = BeautifulSoup(res.text,"lxml")

    return soup

## Function for searching a report-webpage on PTIL (as 'soup') and extracting the url for the pdf




"""
def find_pdf_url_on_webpage(soup):
    report_link = "/"

    ## Find link to the report. In case of two 'similar' tilsynsrapporter, it will always choose the last report
    ## It will distinguish between tilsynrapport and other pdfs, ie. brev
    links = soup.findAll('a', attrs={"class":"d-flex w-100 justify-content-between"})
    #print(links,'\n')
    for link in links:
        pdf_types = link.findAll('h3')
        for pdf_type in pdf_types:
            tmp = pdf_type.get_text()
            #print(tmp, '\n')
            print("Getting the report link...")
            if "Rapport" in tmp or "rapport" in tmp or "tilsynsrapport" in tmp or "Tilsynsrapport" in tmp:
                report_link = link.get('href')
                print("The report link has been obtained. Proceeding")
            else:
                print("There was an error getting the report link. Are you sure that the provided link contains a PDF-file?")
                print(report_link)
            return report_link            
    return "https://www.ptil.no/" + report_link
## Function for opening, reading and converting a pdf url to txt format
"""


def get_all_pdf_link_on_url(pages):
    url = "https://www.ptil.no/"    
    url_reports_pages = "https://www.ptil.no/tilsyn/tilsynsrapporter/?p=" 
    url_to_reports = []    
    url = pages
    response = requests.get(url) #need headers?
    soup = BeautifulSoup(response.content, "html.parser")

    if response.status_code == 200:
        for link in soup.find_all("a", {"class":"pcard"}, href=True):
            try:
                url_to_reports.append(link['href'])
                    
            except AttributeError:
                print("An AttributeError error occured. Proceeding")
                pass
    else:
            print("An error has occured, and it is most likely because the Internt connection")
    print(url_to_reports)
    return url_to_reports

def find_url_to_all_reportpages():
    # Needed for the header in request 
    headers = {"User-Agent": "Mozilla/5.0"} 
    pages = []
    url_reports = []

    # i=0 telles the for-loop to start looking at page 0 in the URL that is being given in the for-loop
    i=0
    end_page = 100
    #this for-loop looks for the pagination in the url, and uses BeautifulSoup to retrieve the page numbers
    # the range(1,10) can be increased to include all pages, or just the pages that are of interest.
    print("Getting the number of pages in the URL. This will take some time based on the Internet connection")
    for i in range(1,end_page):

        #the URL that is being used to retrieve all paginated pages
        url = "https://www.ptil.no/tilsyn/tilsynsrapporter/?p="

        #as can be seen in the URL above, no page number is provided, hence "/?p=". The page number is being assigned in the line below
        url = url + str(i)
        #the URL from the previous step is used to .get all the content on the website
        response = requests.get(url, headers=headers)
        #using the built-in html parser, the content from the response is translated to text?  
        soup = BeautifulSoup(response.content,"html.parser")
            #based on the webpage structure, it can be found that the currect pagination page can be found in the class "page-item active" and "li".
            #This is then transformed to .text.

            # Ensuring that if the response code is 200, else it will fail 
        if response.status_code == 200:
            #try: except: to catch the error that will be presenting if soup_page_number = soup.find().text fails
            try: 
                soup_page_number = soup.find("li", {"class":"page-item active"}).text

                #translated to integer
                soup_page_number = int(soup_page_number) # det er dette som tilsvarer i?

                #print("Appending the current page,", i ,", to the list \"pages\"")
                #pages.append(soup_page_number)
                
                #increases i
                i+=1
            #in this case, an AttributeError will be caught in soup_page_number. Most likely it is becasue there are no more pages to load.
            # An AttributeError is risen when there are no more pages in the pagination to load   
            except AttributeError: 
                print("End of page reached. Last page is", pages[-1])
                #breaking out of the for-loop when there are no more pages to load.
                break
        else:
            print("An error with the URL casued the program to crash. It can either be the website or the the connection to the website. The server responded with error code", 
                    response.status_code)
            break
        #url_reports.append(find_pdf_url_on_webpage(url))
        print("The URL we are visiting now is", url)

        #callinf the function to retrieve all report-url on each page
        get_all_pdf_link_on_url(url)

    return pages
find_url_to_all_reportpages()



def convert_pdf_to_txt(pdf_url):
    
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    f = urlopen(pdf_url).read()
    fp = BytesIO(f)
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos=set()

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
        interpreter.process_page(page)

    pdf_as_text = retstr.getvalue()

    fp.close()
    device.close()
    retstr.close()
    
    return pdf_as_text

## Function for finding participants in revision team
def find_participants_in_revision(idx, report_text):
    
    idx += 1
    
    while report_text[idx] != " " and report_text[idx] != "":
        participants_in_revision = ""
        participants_in_revision += report_text[idx]
        idx += 1
        
    return participants_in_revision

## Function for finding task leader
def find_taskleader(idx, report_text):
    
    task_leader = ""
    idx += 1
    task_leader += report_text[idx]
    
    return task_leader

## Function for finding report title
def find_report_title(idx, report_text): 
    
    idx += 1
    
    while report_text[idx] != " " and report_text[idx] != "":
        report_title = ""
        report_title += report_text[idx]
        idx += 1
        
    return report_title

## Function for finding activity number
def find_activity_number(idx, report_text):
    
    idx += 1
    activity_number = ""
    activity_number += report_text[idx]
    
    return activity_number

## Function for finding date of the report
def find_date(idx, report_text):
    
    idx += 1
    report_date = ""
   
    report_date += report_text[idx]
   
    return report_date

## Function for finding the introduction text 
def find_introduction(idx, report_text):
    
    idx += 1
    report_intro = ""
    
    while not re.compile('2\s+').match(report_text[idx]):
        
        report_intro += report_text[idx]
        idx +=1
        
    return report_intro

## Function for finding installation type
def find_installation_and_type(report_intro):
       
    for key in Installations:
        if key in report_intro:
            installation_type = Installations.get(key)
            installation_name = key
            
            found = True
            break
        
    return installation_name, installation_type 

## Function for looping through pdf and searching for keywords
def find_relevant_info_in_pdf(report_as_a_list_of_sentences):
    
    for idx, line in enumerate(report_as_a_list_of_sentences):
        if ("Deltakere i revisjonslaget" in line) or ("Deltakarar i revisjonslaget" in line):
            participants_in_revision = find_participants_in_revision(idx, report_as_a_list_of_sentences)
            
        if ("Oppgaveleder" in line) or ("Oppgåveleiar" in line):
            taskleader = find_taskleader(idx, report_as_a_list_of_sentences)
            
        if "Aktivitetsnummer" in line:
            activity_number = find_activity_number(idx, report_as_a_list_of_sentences)
            
        if "Dato" in line:
            date = find_date(idx, report_as_a_list_of_sentences)
            
        if "Rapporttittel" in line:
            title = find_report_title(idx, report_as_a_list_of_sentences)
        
        if bool(re.compile('1\s+').match(line)):
            intro = find_introduction(idx, report_as_a_list_of_sentences)
            installation_name, installation_type = find_installation_and_type(intro)
        
    return participants_in_revision, taskleader, activity_number, date, title, installation_name, installation_type

## Function for finding the title of a deviation
def find_deviation_title(deviation):
    
    dev_title = deviation.h3.get_text()
    
    return dev_title

def find_deviation_text(deviation):
    
    dev_text = deviation.p.get_text()
    
    return dev_text

## Function for finding the regulations related to a deviation
def find_deviation_regulations(deviation):
    
    dev_regulations = deviation.find_all('a')
    dev_regulation_list = ""
    
    for dev_regulation in dev_regulations:
        dev_regulation_list += dev_regulation.get_text()
        dev_regulation_list += "\n"
        
    return dev_regulation_list

## Function for finding the title of an improvement point
def find_improvement_title(improvement):
    
    imp_title = improvement.h3.get_text()
    
    return imp_title

## Function for finding the 'reason' of an improvement point
def find_improvement_text(improvement):
    
    imp_text = improvement.p.get_text()
    
    return imp_text

## Function for finding the regulations related to an improvement points
def find_improvement_regulations(improvement):
    
    imp_regulations = improvement.find_all('a')
    imp_regulation_list = ""
    
    for imp_regulation in imp_regulations:
        imp_regulation_list += imp_regulation.get_text()
        imp_regulation_list += "\n"
        
    return imp_regulation_list

## Function for extracting all the deviations and improvement points from the webpage of a report
def find_relevant_info_on_web(webpage_as_soup, report):
    
    deviations = webpage_as_soup.find_all('div', attrs={"class":"tab-pane","id":re.compile('deviation.*')})
    dev_cntr = 0
    
    #print("Here are the deviations:")
    #print("")
    for deviation in deviations:
        dev_title = find_deviation_title(deviation)
        dev_text = find_deviation_text(deviation)
        dev_regulations = find_deviation_regulations(deviation)
        dev_cntr += 1
        
        new_deviation = Deviation(dev_title, dev_text, dev_regulations)
        
        report.deviation_list.append(new_deviation)
    
    #print("Number of deviations found: ", dev_cntr)
    
    improvements = webpage_as_soup.find_all('div', attrs={"class":"tab-pane","id":re.compile('improvementPoint.*')})
    imp_cntr = 0
    #print("Here are the improvement points:")
    #print("")
    for improvement in improvements:
        imp_title = find_improvement_title(improvement)
        imp_text = find_improvement_text(improvement)
        imp_regulations = find_improvement_regulations(improvement)
        imp_cntr += 1
        new_improvement = Improvementpoint(imp_title, imp_text, imp_regulations)
        #print("Improvement point title:")
        #print(new_improvement.title)
        #print("")
        #print("Improvement point explanatory text:")
        #print(new_improvement.description)
        #print("")
        #print("All regulations:")
        #print(new_improvement.regulations)
        #print("----")
    
        report.improvement_list.append(new_improvement)
    
    #print("Number of improvement points found: ", imp_cntr)
    #print("")
    

"""
Functions for writing the excel sheet
"""

   ## Function for opening the excel sheet one wishes to write to
def open_excel_sheet(path):
    
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    return wb, ws

## Function for finding the next row for filling in a finding
def find_next_row(ws):
    
    for next_row, row in enumerate(ws['B'], 1):
        if (row.value == None) and (next_row > 17):
            break
            
    if len(ws['B']) == next_row:
        next_row += 1
    
    return next_row

## Function for filling in all the findings from all the new reports
def fill_in_database(filename, list_of_reports):
    
    wb, ws = open_excel_sheet(filename)
    next_row = find_next_row(ws)
    
    for report in list_of_reports:
        
        tilsyn_id =  ws[next_row - 1][2].value + 1
        
        for deviation in report.deviation_list:
            
            finding_id = ws[next_row - 1][1].value + 1
            ws.cell(column = 2, row = next_row, value = finding_id)
            ws.cell(column = 3, row = next_row, value = tilsyn_id)
            ws.cell(column = 5, row = next_row, value = finding_id)
            ws.cell(column = 8, row = next_row, value = report.installation_name)
            ws.cell(column = 9, row = next_row, value = report.installation_type)
            #ws.cell(column = 10, row = next_row, value = report.authority)
            ws.cell(column = 11, row = next_row, value = report.year)
            ws.cell(column = 12, row = next_row, value = report.date)
            ws.cell(column = 13, row = next_row, value = report.activity_number)
            ws.cell(column = 14, row = next_row, value = report.title)
            ws.cell(column = 16, row = next_row, value = report.taskleader)
            ws.cell(column = 17, row = next_row, value = report.participants_in_revision)
            ws.cell(column = 18, row = next_row, value = report.count_participants)
            ws.cell(column = 19, row = next_row, value = report.url)
            ws.cell(column = 21, row = next_row, value = len(report.deviation_list))
            ws.cell(column = 22, row = next_row, value = len(report.improvement_list))
            
            ws.cell(column = 24, row = next_row, value = deviation.finding_type)
            ws.cell(column = 25, row = next_row, value = deviation.title)
            ws.cell(column = 26, row = next_row, value = deviation.description)
            ws.cell(column = 27, row = next_row, value = deviation.regulations)
            #ws.cell(column = 29, row = next_row, value = deviation.kategori)
            
            next_row += 1
            
        for improvement in report.improvement_list:
            
            finding_id = ws[next_row - 1][1].value + 1
            ws.cell(column = 2, row = next_row, value = finding_id)
            ws.cell(column = 3, row = next_row, value = tilsyn_id)
            ws.cell(column = 5, row = next_row, value = finding_id)
            ws.cell(column = 8, row = next_row, value = report.installation_name)
            ws.cell(column = 9, row = next_row, value = report.installation_type)
            #ws.cell(column = 10, row = next_row, value = report.authority)
            ws.cell(column = 11, row = next_row, value = report.year)
            ws.cell(column = 12, row = next_row, value = report.date)
            ws.cell(column = 13, row = next_row, value = report.activity_number)
            ws.cell(column = 14, row = next_row, value = report.title)
            ws.cell(column = 16, row = next_row, value = report.taskleader)
            ws.cell(column = 17, row = next_row, value = report.participants_in_revision)
            ws.cell(column = 18, row = next_row, value = report.count_participants)
            ws.cell(column = 19, row = next_row, value = report.url)
            ws.cell(column = 21, row = next_row, value = len(report.deviation_list))
            ws.cell(column = 22, row = next_row, value = len(report.improvement_list))
            
            ws.cell(column = 24, row = next_row, value = improvement.finding_type)
            ws.cell(column = 25, row = next_row, value = improvement.title)
            ws.cell(column = 26, row = next_row, value = improvement.description)
            ws.cell(column = 27, row = next_row, value = improvement.regulations)
            #ws.cell(column = 29, row = next_row, value = improvement.kategori)
            
            next_row += 1
            
    wb.save(filename)

    """Horrible main() that could be cleaned up"""

def main():
    print("Running main()")    
    ## INSERT HERE:Function() for going through 'https://www.ptil.no/tilsyn/tilsynsrapporter//GetData?pageindex=0&pagesize=10000'
    ## -webpage, and making a list of all the urls for new reports that are not in the excel database yet
    
    ## list_of_webpage_urls_for_reports = []
    
    ## Make list for all the reports
    report_list = []
      
    ## for-loop through all the reports in the list
        ## Everything here should be in the for-loop
        
        ## At every iteration, we extract the pdf url, corresponding to the webpage url


    # Kjørere script for å hente alle rapportene
    find_url_to_all_reportpages()
    #pages = pages.find_url_to_all_reportpages()

    test_url = "https://www.ptil.no/tilsyn/tilsynsrapporter/2019/conocophillips-ekofisk-stimuleringsoperasjon-fra-fartoy/"
    
    url_soup = make_soup(test_url)

    pdf_link = find_pdf_url_on_webpage(url_soup)
    pdf_link = "https://www.ptil.no/" + pdf_link
    #print(pdf_link)

        ## Get pdf as txt
    pdfText = convert_pdf_to_txt(pdf_link)

        ## Split on "\n"
    pdf_as_list_of_words = pdfText.split("\n")
    
        ## Looping through report and searching for keywords (returned as string)
    participants_in_revision, taskleader, activity_number, report_date, report_title, installation_name, installation_type = find_relevant_info_in_pdf(pdf_as_list_of_words)
    
        ## Make a report object
    report = Report(report_date.strip(' ')[-4:], report_date, activity_number, report_title, taskleader,
                   participants_in_revision, participants_in_revision.count(",") + participants_in_revision.count(" og ") + 1,
                   pdf_link, installation_name, installation_type)
    
        ## Find deviations and improvement points, make objects then add them to the report
    find_relevant_info_on_web(url_soup, report)
    
        ## INSERT HERE: Probably a function that takes in all the deviations, improvement points, as well as a ML model of
        ## choice, then predicts the cathegories. Maybe require human assistance if classification % is less than a threshold
        
        ## Append new finding to list
    report_list.append(report)
    
    ## Write everything to file
    ## NB!!! FOR SOME REASON, IT IS NOT POSSIBLE TO WRITE TO ORIGNAL TILSYSRAPPORT WE WERE PROVIDED(?) THE FILE GETS CORRUPCTED.
    ## WE HAD TO MAKE A SIMPLIFIED VERSION, WHERE WRITING WORKED FINE!
    fill_in_database((r'C:\\Users\\HANNORU\source\\Machine learning summer project\\Tilsynsdatabase_revised_simplified.xlsx'), report_list)
    print("Successfully written database")
 
    ## Printing the extracted information from the report
    print("URL for pdf:")
    print(report.url)
    print ("-----------------------")
    print ("Aktivitetsnummer er følgende:")
    print (report.activity_number)
    print ("")
    print ("RapportTittel er følgende:")
    print (report.title)
    print ("")
    print ("RapportDato er følgende:")
    print (report.date)
    print ("")
    print ("Oppgaveleder er følgende:")
    print (report.taskleader)
    print ("")
    print ("Deltakere i revisjonslaget er følgende:")
    print (report.participants_in_revision)
    print("")
    print("Navnet på installasjonen:")
    print(report.installation_name)
    print("")
    print("Type:")
    print(report.installation_type)
    print("------------------------")
    
    for test_dev in report.deviation_list:
        print("Tittel på avvik:")
        print(test_dev.title)
        print("")
        print("Avvikets beskrivende tekst:")
        print(test_dev.description)
        print("")
        print("Alle regelhenvisninger:")
        print(test_dev.regulations)
        print("----")
        
    for test_imp in report.improvement_list:
        print("Tittel på forbedringspunkt:")
        print(test_imp.title)
        print("")
        print("Avvikets beskrivende tekst:")
        print(test_imp.description)
        print("")
        print("Alle regelhenvisninger:")
        print(test_imp.regulations)
        print("----")
  

main()   


### THIS SHOULD PROB BE IN IT'S OWN FILE
## Dictionary with installations and their type (alphabetical order). Installations registered with more than one type
## in the database is not included yet and temporarily commented out. They will be included when we know what type they
## belong to.

